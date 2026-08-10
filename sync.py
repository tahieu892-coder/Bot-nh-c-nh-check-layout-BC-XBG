"""Đồng bộ danh sách BC – AM từ file bên ngoài.

Hai nguồn được hỗ trợ:
  1. Google Sheets (hoặc bất kỳ URL trả về CSV/XLSX) — đặt biến SHEET_URL, bot tự kéo về.
  2. File .csv / .xlsx gửi thẳng vào chat Telegram kèm caption /capnhat.

Bot tự dò dòng tiêu đề trong 10 dòng đầu và nhận 4 cột theo tên:
  Mã BC | Bưu Cục | Họ Tên (AM) | Tele (@username của AM)
Các cột trùng tên xuất hiện sau đó (vd cột "AM"/"Tele AM" dùng làm danh sách tham chiếu)
sẽ bị bỏ qua — mỗi vai trò chỉ lấy cột khớp ĐẦU TIÊN từ trái sang.
"""

import csv
import io
import os
import re
import unicodedata

import httpx

import db

# Các dòng trong sheet KHÔNG phải bưu cục cần chụp layout (vd Kho Chuyển Tiếp).
BO_QUA_BC = {x.strip().upper() for x in os.getenv("BO_QUA_BC", "").split(",") if x.strip()}
BO_QUA_TU_KHOA = [x.strip() for x in
                  os.getenv("BO_QUA_TU_KHOA", "kho chuyen tiep,kct").split(",") if x.strip()]

# Từ khoá nhận diện cột (đã bỏ dấu, viết thường). Thứ tự trong tuple = độ ưu tiên.
VAI_TRO = {
    "code": ("ma bc", "mabc", "ma buu cuc", "ma bưu cuc", "code", "ma"),
    "name": ("buu cuc", "ten bc", "ten buu cuc", "tenbc", "name"),
    "am_name": ("ho ten", "ho va ten", "ten am", "am phu trach", "nguoi phu trach",
                "quan ly", "am"),
    "am_tele": ("tele", "telegram", "tele am", "nick tele", "username"),
}


def _bo_dau(s) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s.replace("đ", "d").replace("Đ", "D")).strip().lower()


def _o(v) -> str:
    """Chuẩn hoá 1 ô: số trong Excel hay ra dạng 23009000.0 → cắt phần .0."""
    if v is None:
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def _khop(ten_cot: str, tu_khoa: tuple) -> bool:
    return any(ten_cot == k for k in tu_khoa)


def _map_cot(header: list) -> dict:
    """Gán cột cho từng vai trò, quét trái→phải, mỗi vai trò lấy cột khớp đầu tiên."""
    h = [_bo_dau(x) for x in header]
    cols: dict[str, int] = {}
    for i, ten_cot in enumerate(h):
        if not ten_cot:
            continue
        for vai_tro, tu_khoa in VAI_TRO.items():
            if vai_tro not in cols and _khop(ten_cot, tu_khoa):
                cols[vai_tro] = i
                break
    return cols


def parse_rows(rows: list[list]) -> list[tuple[str, str, str, str]]:
    """Đọc bảng thô thành danh sách (mã BC, tên BC, tên AM, @tele AM)."""
    rows = [r for r in rows if any(_o(c) for c in r)]
    if not rows:
        return []

    # Dò dòng tiêu đề: dòng đầu tiên trong 10 dòng đầu có cột mã BC.
    cols, bat_dau = None, 0
    for i, r in enumerate(rows[:10]):
        m = _map_cot(r)
        if "code" in m:
            cols, bat_dau = m, i + 1
            break
    if cols is None:  # file không có tiêu đề → dùng thứ tự cột mặc định
        cols, bat_dau = {"code": 0, "name": 1, "am_name": 2, "am_tele": 3}, 0

    out = []
    for r in rows[bat_dau:]:
        def lay(vai_tro):
            i = cols.get(vai_tro, -1)
            return _o(r[i]) if 0 <= i < len(r) else ""

        ma = lay("code").upper()
        if not ma or not re.search(r"[A-Z0-9]", ma):
            continue
        out.append((ma, lay("name"), lay("am_name"), lay("am_tele").lstrip("@").strip()))
    return out


def from_csv_bytes(data: bytes) -> list[tuple[str, str, str, str]]:
    text = data.decode("utf-8-sig", errors="replace")
    delim = ";" if text.count(";") > text.count(",") else ","
    return parse_rows([r for r in csv.reader(io.StringIO(text), delimiter=delim)])


def from_xlsx_bytes(data: bytes) -> list[tuple[str, str, str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return parse_rows([list(r) for r in ws.iter_rows(values_only=True)])


def csv_url(url: str) -> str:
    """Đổi link Google Sheets thường thành link xuất CSV của đúng tab đang mở."""
    m = re.search(r"docs\.google\.com/spreadsheets/d/([\w-]+)", url)
    if not m:
        return url
    gid = re.search(r"[#&?]gid=(\d+)", url)
    return (f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=csv"
            f"&gid={gid.group(1) if gid else '0'}")


async def from_url(url: str) -> list[tuple[str, str, str, str]]:
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        r = await client.get(csv_url(url))
        r.raise_for_status()
    if url.lower().endswith(".xlsx") or "spreadsheetml" in r.headers.get("content-type", ""):
        return from_xlsx_bytes(r.content)
    return from_csv_bytes(r.content)


def loc_bo_qua(records: list[tuple[str, str, str, str]]):
    """Tách các dòng không phải BC cần chụp layout (Kho Chuyển Tiếp, mã trong BO_QUA_BC)."""
    giu, bo = [], []
    tu_khoa = [_bo_dau(k) for k in BO_QUA_TU_KHOA]
    for r in records:
        ma, ten, am, _ = r
        hay = _bo_dau(f"{ten} {am}")
        if ma in BO_QUA_BC or any(k and k in hay for k in tu_khoa):
            bo.append(r)
        else:
            giu.append(r)
    return giu, bo


def apply(records: list[tuple[str, str, str, str]], deactivate: bool = True) -> dict:
    """Ghi danh sách mới vào DB. Trả về tóm tắt thay đổi để báo cho quản lý."""
    if not records:
        return {"loi": "File rỗng hoặc không tìm thấy cột 'Mã BC'."}

    records, bo_qua = loc_bo_qua(records)
    if not records:
        return {"loi": "Sau khi lọc thì không còn BC nào. Kiểm tra lại BO_QUA_TU_KHOA."}

    hien_tai = {r["code"]: r for r in db.list_bc(active_only=True)}
    moi, doi_am, giu = [], [], 0
    canh_bao = None

    # Chặn tai nạn: file lỗi/mất dòng làm bot xoá gần hết danh sách đang theo dõi.
    if deactivate and hien_tai and len(records) < len(hien_tai) * 0.5:
        deactivate = False
        canh_bao = (f"File chỉ có {len(records)} BC trong khi đang theo dõi {len(hien_tai)}. "
                    f"Bot đã <b>bỏ qua bước ngừng theo dõi</b> để tránh xoá nhầm — "
                    f"kiểm tra lại file, hoặc dùng <code>/xoabc</code> để gỡ thủ công.")

    for ma, ten, am, tele in records:
        cu = hien_tai.get(ma)
        if cu is None:
            moi.append((ma, ten, am))
        elif (cu["am_name"] or "") != am:
            doi_am.append((ma, cu["am_name"] or "-", am or "-"))
        else:
            giu += 1
        db.upsert_bc(ma, ten, am)
        if am and tele:
            db.upsert_am(am, username=tele)  # lưu @username để tag AM khi nhắc

    ma_moi = {r[0] for r in records}
    ngung = []
    if deactivate:
        for ma, r in hien_tai.items():
            if ma not in ma_moi:
                db.remove_bc(ma)
                ngung.append((ma, r["name"] or ""))

    return {"tong": len(records), "moi": moi, "doi_am": doi_am, "giu": giu,
            "ngung": ngung, "canh_bao": canh_bao, "bo_qua": bo_qua,
            "thieu_tele": sorted({am for _, _, am, tele in records if am and not tele})}


def tom_tat(kq: dict) -> str:
    if kq.get("loi"):
        return f"❌ Đồng bộ thất bại: {kq['loi']}"

    out = ["🔄 <b>ĐÃ ĐỒNG BỘ DANH SÁCH BC</b>",
           f"Theo dõi: <b>{kq['tong']}</b> BC · Giữ nguyên: {kq['giu']}"
           + (f" · Bỏ qua: {len(kq['bo_qua'])}" if kq.get("bo_qua") else "")]
    if kq["moi"]:
        out.append(f"\n➕ <b>Thêm mới ({len(kq['moi'])})</b>")
        out += [f"• <code>{ma}</code> {ten} — AM: {am or '-'}" for ma, ten, am in kq["moi"][:30]]
    if kq["doi_am"]:
        out.append(f"\n🔁 <b>Đổi AM ({len(kq['doi_am'])})</b>")
        out += [f"• <code>{ma}</code>: {cu} → <b>{new}</b>" for ma, cu, new in kq["doi_am"][:30]]
    if kq["ngung"]:
        out.append(f"\n➖ <b>Ngừng theo dõi ({len(kq['ngung'])})</b>")
        out += [f"• <code>{ma}</code> {ten}" for ma, ten in kq["ngung"][:30]]
    if not (kq["moi"] or kq["doi_am"] or kq["ngung"]):
        out.append("\nKhông có thay đổi nào.")
    if kq.get("thieu_tele"):
        out.append(f"\n⚠️ Chưa có nick Telegram, bot không tag được: "
                   f"{', '.join(kq['thieu_tele'][:15])}")
    if kq.get("canh_bao"):
        out.append(f"\n🛑 {kq['canh_bao']}")
    return "\n".join(out)
